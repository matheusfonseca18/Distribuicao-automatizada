import json
import os
import random
import pandas as pd
from datetime import datetime, timedelta
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.worksheet.table import Table, TableStyleInfo
from CTkMessagebox import CTkMessagebox
from tkinter import filedialog
from itertools import cycle
from apelidos_utils import carregar_apelidos, montar_mapa_apelidos, nome_exibicao


# lê os arquivos JSON
def ler_arquivos_json(nome_arquivo):
    if os.path.exists(nome_arquivo):
        with open(nome_arquivo, "r") as f:
            return json.load(f)
    return {}

def salvar_arquivo_json(nome_arquivo, dados):
    if os.path.exists(nome_arquivo):
        with open(nome_arquivo, "w") as f:
            json.dump(dados, f, indent=4)

def converter_data(data_str):
    data_limpa = str(data_str).split(" ")[0]
    formatos = ("%d/%m/%Y", "%Y-%m-%d")

    for formato in formatos:
        try:
            return datetime.strptime(data_limpa, formato)
        except ValueError:
            pass

    try:
        return datetime.fromisoformat(str(data_str))
    except ValueError:
        return None

def datas_consecutivas(data_anterior, data_atual):
    anterior = converter_data(data_anterior)
    atual = converter_data(data_atual)
    return bool(anterior and atual and anterior + timedelta(days=1) == atual)

def formatar_data_br(data_str):
    data = converter_data(data_str)
    if data:
        return data.strftime("%d/%m/%Y")
    return str(data_str).split(" ")[0]

def texto_formula_excel(valor):
    return str(valor).replace('"', '""')

def formatar_equipe(equipe, mapa_apelidos):
    return " / ".join(nome_exibicao(nome, mapa_apelidos) for nome in equipe)

def distribuir_por_turno(lista_colaboradores, lista_atividades, historico, data_atual, turno_alvo):
    atividades_ordenadas = sorted(lista_atividades, key=lambda x: int(x['grupo']))
    colaboradores_turno = list(dict.fromkeys(lista_colaboradores))
    equipe_reduzida = len(colaboradores_turno) <= 4
    disponiveis = list(colaboradores_turno)
    ordem_original = {colab: idx for idx, colab in enumerate(colaboradores_turno)}
    alocacao = {atv['nome']: [] for atv in lista_atividades}
    mapa_atividades = {atv['nome']: atv for atv in lista_atividades}
    mapa_links = {atv['nome']: [l.strip() for l in atv.get('link', "").split(",") if l.strip()] for atv in lista_atividades}
    data_atual_formatada = formatar_data_br(data_atual)

    dias_trabalhados = historico.setdefault("__dias_trabalhados__", {})
    for colaborador in colaboradores_turno:
        registro_dias = dias_trabalhados.setdefault(colaborador, {"qtd": 0, "ultima_data": ""})
        if registro_dias.get("ultima_data") != data_atual_formatada:
            registro_dias["qtd"] = int(registro_dias.get("qtd", 0)) + 1
            registro_dias["ultima_data"] = data_atual_formatada

    def qtd_por_turno(atividade):
        qtd_configurada = int(atividade.get("pessoas", {}).get(turno_alvo, 0))
        if equipe_reduzida and qtd_configurada > 0:
            return 1
        return qtd_configurada

    def pode_exceder(nome_atv):
        atividade = mapa_atividades.get(nome_atv, {})
        return atividade.get('pode_exceder') == "S"

    def tem_vaga_ou_excede(nome_atv):
        atividade = mapa_atividades.get(nome_atv)
        if not atividade:
            return False

        return len(alocacao[nome_atv]) < qtd_por_turno(atividade) or pode_exceder(nome_atv)

    def atividades_previstas(nome_atv):
        previstas = [nome_atv]

        for nome_linkado in mapa_links.get(nome_atv, []):
            if nome_linkado in alocacao and tem_vaga_ou_excede(nome_linkado):
                previstas.append(nome_linkado)

        return list(dict.fromkeys(previstas))

    def registrar_alocacao(nome_atv, colaborador, consumir_disponivel=False):
        if colaborador in alocacao[nome_atv]:
            return False

        alocacao[nome_atv].append(colaborador)

        if consumir_disponivel and colaborador in disponiveis:
            disponiveis.remove(colaborador)

        if colaborador not in historico:
            historico[colaborador] = {}
        if nome_atv not in historico[colaborador]:
            historico[colaborador][nome_atv] = {"qtd": 0, "ultima_data": ""}

        historico[colaborador][nome_atv]["qtd"] = int(historico[colaborador][nome_atv].get("qtd", 0)) + 1
        historico[colaborador][nome_atv]["ultima_data"] = data_atual_formatada

        return True

    def calcular_peso_colaborador(colab, atividades_do_peso):
        historico_colab = historico.get(colab, {})
        qtd_dias = max(1, int(dias_trabalhados.get(colab, {}).get("qtd", 1)))

        total_geral = sum(info.get("qtd", 0) for info in historico_colab.values())
        total_atividade = sum(
            historico_colab.get(atv, {"qtd": 0}).get("qtd", 0)
            for atv in atividades_do_peso
        )

        repetiu_ontem = any(
            datas_consecutivas(
                historico_colab.get(atv, {"ultima_data": ""}).get("ultima_data", ""),
                data_atual
            )
            for atv in atividades_do_peso
        )

        return (
            1 if repetiu_ontem else 0,
            total_atividade / qtd_dias,
            total_geral / qtd_dias,
            total_atividade,
            total_geral,
            ordem_original.get(colab, 0)
        )

    def pode_receber_atividade(colab, atividades_do_peso):
        for nome_previsto in atividades_do_peso:
            if pode_exceder(nome_previsto):
                continue

            ultima_data = historico.get(colab, {}).get(nome_previsto, {}).get("ultima_data", "")
            if datas_consecutivas(ultima_data, data_atual_formatada):
                return False

        return True

    def atribuir_colaborador(atividade, colaborador):
        nome_atv = atividade['nome']
        registrar_alocacao(nome_atv, colaborador, consumir_disponivel=True)

        for nome_linkado in mapa_links.get(nome_atv, []):
            if nome_linkado in alocacao and tem_vaga_ou_excede(nome_linkado):
                registrar_alocacao(nome_linkado, colaborador)

    for atividade in atividades_ordenadas:
        nome_atv = atividade['nome']

        while len(alocacao[nome_atv]) < qtd_por_turno(atividade):
            previstas = atividades_previstas(nome_atv)
            candidatos = [
                colab for colab in disponiveis
                if colab not in alocacao[nome_atv] and pode_receber_atividade(colab, previstas)
            ]

            if not candidatos:
                break

            candidatos.sort(key=lambda colab: calcular_peso_colaborador(colab, previstas))
            atribuir_colaborador(atividade, candidatos[0])

        alocacao[nome_atv] = list(dict.fromkeys(alocacao[nome_atv]))

    # Distribui quem sobrou
    if disponiveis:
        for atividade in atividades_ordenadas:
            if atividade.get('pode_exceder') == "S":
                nome_atv_excede = atividade['nome']
                candidatos = list(disponiveis)
                candidatos.sort(key=lambda colab: calcular_peso_colaborador(colab, [nome_atv_excede]))

                for colaborador in candidatos:
                    atribuir_colaborador(atividade, colaborador)

                alocacao[nome_atv_excede] = list(dict.fromkeys(alocacao[nome_atv_excede]))
                break

    return alocacao, historico


def gerar_distribuicao():
    # Carrega os dados
    escala_json = ler_arquivos_json("dados_escala.json")
    atividades = ler_arquivos_json("atividades_pendentes.json")
    mapa_apelidos = montar_mapa_apelidos(carregar_apelidos())
    # Carrega o histórico existente ou cria um vazio
    historico_geral = ler_arquivos_json("historico.json") 

    resultado_final = {}

    for data_str in sorted(escala_json.keys(), key=lambda data: converter_data(data) or datetime.max):
        turnos = escala_json[data_str]
        data_formatada = formatar_data_br(data_str)
        resultado_final[data_formatada] = {}
        
        for nome_do_turno, lista_colaboradores in turnos.items():
            # recebe a alocação e o histórico atualizado
            alocacao_turno, historico_geral = distribuir_por_turno(
                lista_colaboradores, 
                atividades, 
                historico_geral,
                data_formatada,
                nome_do_turno
            )
            resultado_final[data_formatada][nome_do_turno] = alocacao_turno

    # GERAÇÃO DO EXCEL
    caminho_salvar = filedialog.asksaveasfilename(
        defaultextension='.xlsx', 
        filetypes=[("Arquivo Excel", "*.xlsx")], 
        title='Salvar Distribuição', 
        initialfile=f'Distribuição.xlsx'
    )

    if caminho_salvar:
        try:
            wb = Workbook()
            
            # ABA DISTRIBUIÇÃO (DADOS DO PANDAS)
            ws_dist = wb.active
            ws_dist.title = "Distribuição"
            
            # Criar o DataFrame df_final
            linhas_base = []
            lista_turnos = sorted(list(set(t for turnos in escala_json.values() for t in turnos.keys())))
            for atividade in atividades:
                nome_atv_fixo = atividade['nome'].strip().upper()
                for turno in lista_turnos:
                    linhas_base.append({"ATIVIDADE": nome_atv_fixo, "TURNO": turno})
            
            df_final = pd.DataFrame(linhas_base)
            for data, turnos_da_data in resultado_final.items():
                data_curta = formatar_data_br(data)
                coluna_colaboradores = []
                for index, row in df_final.iterrows():
                    atv_procurada = row["ATIVIDADE"]
                    turno_procurado = row["TURNO"]
                    alocacao_turno = turnos_da_data.get(turno_procurado, {})
                    equipe = next((lista for n, lista in alocacao_turno.items() if n.strip().upper() == atv_procurada), [])
                    coluna_colaboradores.append(formatar_equipe(equipe, mapa_apelidos))
                df_final[data_curta] = coluna_colaboradores

            # Converter DataFrame do Pandas para linhas do Openpyxl
            for r in dataframe_to_rows(df_final, index=False, header=True):
                ws_dist.append(r)

            # ABA HISTÓRICO
            ws_historico = wb.create_sheet(title="Histórico")
            
            # Identifica os colaboradores únicos na escala
            todos_colaboradores = set()
            for data_str, turnos in escala_json.items():
                for nome_turno, lista in turnos.items():
                    for nome_colab in lista:
                        todos_colaboradores.add((nome_colab.strip().upper(), nome_turno.upper()))
            
            # Ordenar por turno e depois nome
            lista_colaboradores_ordenada = sorted(list(todos_colaboradores), key=lambda x: (x[1], x[0]))

            # Identifica as atividades para criar as colunas
            nomes_atividades = [atv['nome'].strip().upper() for atv in atividades]
            colunas_historico = ['COLABORADOR', 'TURNO'] + nomes_atividades
            ws_historico.append(colunas_historico)

            # Preenche os dados e as fórmulas
            for i, (nome_colab, turno_colab) in enumerate(lista_colaboradores_ordenada, start=2):
                # Escreve Nome e Turno
                ws_historico.cell(row=i, column=1).value = nome_colab
                ws_historico.cell(row=i, column=2).value = turno_colab
                nome_colab_busca = texto_formula_excel(nome_exibicao(nome_colab, mapa_apelidos))
                turno_colab_formula = texto_formula_excel(turno_colab)
                
                # Para cada atividade (coluna), insere a fórmula de contagem
                for j, nome_atv in enumerate(nomes_atividades, start=3):                    
                    nome_atv_formula = texto_formula_excel(nome_atv)
                    formula = (
                        f'=SUMPRODUCT((Distribuição!$A$2:$A$1000="{nome_atv_formula}")*'
                        f'(Distribuição!$B$2:$B$1000="{turno_colab_formula}")*'
                        f'(--ISNUMBER(SEARCH("{nome_colab_busca}", Distribuição!$C$2:$AG$1000))))'
                    )
                    ws_historico.cell(row=i, column=j).value = formula

            # Formata como Tabela
            ultima_col_letra = chr(64 + len(colunas_historico)) # Converte índice para letra
            if len(colunas_historico) > 26: # se tiver muitas atividades
                ultima_col_letra = "Z" 
                
            ref_tabela = f'A1:{ultima_col_letra}{len(lista_colaboradores_ordenada) + 1}'
            tab = Table(displayName='Tabela_Historico', ref=ref_tabela)
            style = TableStyleInfo(name="TableStyleMedium9", showRowStripes=True)
            tab.tableStyleInfo = style
            ws_historico.add_table(tab)

            # ABA RADIO
            ws_radio = wb.create_sheet(title="RADIO")
        
            pracas = [
                "PO", "SÃO PAULO", "PORTO ALEGRE", "RIO DE JANEIRO",
                "BELO HORIZONTE", "CURITIBA", "SALVADOR", "GOIÂNIA",
                "BRASÍLIA", "SANTOS", "FLORIANOPOLIS",
                "VITÓRIA", "JUIZ DE FORA", "CAMPINAS", "RECIFE"
            ]

            turnos_lista = ["MANHA", "TARDE"]

            #Cabeçalho
            datas_ordenadas = sorted(resultado_final.keys(), key=lambda data: converter_data(data) or datetime.max)
            datas_curtas = [formatar_data_br(d) for d in datas_ordenadas]
            ws_radio.append(["PRACA", "TURNO"] + datas_curtas)


            #Linhas base
            linhas_base = []
            for praca in pracas:
                for turno in turnos_lista:
                    linhas_base.append({"PRACA": praca, "TURNO": turno})

            #distribuição
            nomes_organizados = {}
            for data_chave in datas_ordenadas:
                nomes_organizados[data_chave] = {}
                dia_alocacao = resultado_final.get(data_chave, {})
                
                for turno in turnos_lista:
                    dados_radio = dia_alocacao.get(turno, {}).get("RADIO", "")
                    
                    # Limpeza dos nomes
                    if isinstance(dados_radio, list):
                        lista = [nome_exibicao(n.strip(), mapa_apelidos) for n in dados_radio if n.strip()]
                    else:
                        lista = [nome_exibicao(n.strip(), mapa_apelidos) for n in dados_radio.split("/") if n.strip()]
                    
                    # Criamos o distribuidor circular para cada turno deste dia
                    nomes_organizados[data_chave][turno] = cycle(lista) if lista else None

            # 2. Distribuição nas Linhas
            for idx_linha, base in enumerate(linhas_base, start=2):
                praca_atual = base["PRACA"]
                turno_atual = base["TURNO"]

                # Escreve as colunas fixas
                ws_radio.cell(row=idx_linha, column=1).value = praca_atual
                ws_radio.cell(row=idx_linha, column=2).value = turno_atual

                # 3. Preenche as colunas de datas para ESTA linha
                for idx_coluna, data_chave in enumerate(datas_ordenadas, start=3):
                    # Pegamos o distribuidor que já preparamos lá em cima
                    distribuidor = nomes_organizados[data_chave][turno_atual]
                    
                    if distribuidor:
                        valor_final = next(distribuidor)
                    else:
                        valor_final = ""
                        
                    ws_radio.cell(row=idx_linha, column=idx_coluna).value = valor_final

                # Salvar o arquivo final
                wb.save(caminho_salvar)

            CTkMessagebox(title="Sucesso", message="Distribuição baixado com sucesso!", icon="check")
        except Exception as e:
            CTkMessagebox(title="Erro", message=f"Erro ao salvar: {e}", icon="cancel")
    
if __name__ == "__main__":
    gerar_distribuicao()
