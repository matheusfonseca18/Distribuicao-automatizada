import json
import os
import pandas as pd
from datetime import datetime
from CTkMessagebox import CTkMessagebox
from tkinter import filedialog
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

# Validar se a entrada é um número
def validar_numero(digitar, valor):
    if digitar == "1":
        return valor.isdigit()
    return True

def pintar_range(ws, celula, cor):
    fill = PatternFill(start_color=cor, end_color=cor, fill_type='solid')
    for linha in ws[celula]:
        for cel in linha:
            cel.fill = fill

def baixar_escala_modelo():
    ano_atual = datetime.now().year
    datas = pd.date_range(start=f"{ano_atual}-01-01", end=f"{ano_atual}-12-31")
    dias_mes = [d.strftime('%d/%m/%Y') for d in datas]
    dias_semana_map = {
        'Monday': 'Seg', 'Tuesday': 'Ter', 'Wednesday': 'Qua', 
        'Thursday': 'Qui', 'Friday': 'Sex', 'Saturday': 'Sab', 'Sunday': 'Dom'
    }
    dias_semana = [dias_semana_map[d.strftime('%A')] for d in datas]

    header_l1 = ['NOME', 'TURNO', 'GRUPO'] + dias_mes
    header_l2 = ['', '', ''] + dias_semana

    linha_exemplo = ['Colaborador 1', 'MANHÃ', '1'] + (['OK'] * len(dias_mes))
    avisos = ['AVISOS', 'NÃO ALTERE A ORDEM DAS COLUNAS', 'NÃO REMOVA NENHUMA DAS 3 PRIMEIRA COLUNAS (A, B e C)', 'NÃO APAGUE A SEGUNDA LINHA']
    
    dados_finais = [header_l2, linha_exemplo, avisos]

    df_modelo = pd.DataFrame(dados_finais, columns=header_l1)

    caminho_salvar = filedialog.asksaveasfilename(
        defaultextension='.xlsx', 
        filetypes=[("Arquivo Excel", "*.xlsx")], 
        title='Salvar modelo de escala', 
        initialfile=f'Modelo_Escala.xlsx'
    )

    if caminho_salvar:
        try:
            df_modelo.to_excel(caminho_salvar, index=False)

            wb = load_workbook(caminho_salvar)
            ws = wb.active

            # Ajusta largura das colunas
            for col in ws.columns:
                max_length = 0
                col_letter = col[0].column_letter

                for cell in col:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                ws.column_dimensions[col_letter].width = max_length + 2

            pintar_range(ws, 'A1:C1', '8DB4E2') #azul cabeçalho
            pintar_range(ws, 'D1:ND1', 'D9EAF7') #azul claro datas
            pintar_range(ws, 'D2:ND2', 'EBF1DE') #verde claro dias semana
            pintar_range(ws, 'A4:D4', 'FFFF00') #amarelo aviso

            for col in range(4, ws.max_column +1):
                cell = ws.cell(row=2, column=col)
                if isinstance(cell.value, str):
                    texto = cell.value.lower()
                    if 'sab' in texto or 'dom' in texto:
                        cell.fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid') #vermelho fim de semana

            wb.save(caminho_salvar)

            CTkMessagebox(title="Sucesso", message="Modelo baixado com sucesso!", icon="check")
        except Exception as e:
            CTkMessagebox(title="Erro", message=f"Erro ao salvar: {e}", icon="cancel")


# FUNÇÕES para salvar os dados qnd o app for fechado e carregar qnd aberto
def salvar_dados(nome_arquivo, dados_arquivo):
    try:
        with open(nome_arquivo, 'w', encoding='utf-8') as f:
            json.dump(dados_arquivo, f, indent=4)
    except Exception as e:
        CTkMessagebox(title="Erro de Arquivo", 
                      message=f"Não foi possível salvar os dados da atividade no arquivo: {e}",
                      icon="cancel")
        
def carregar_dados(nome_arquivo):
    if os.path.exists(nome_arquivo):
        try:
            with open(nome_arquivo, 'r', encoding='utf-8') as f:
                return json.load(f)
        except Exception as e:
            CTkMessagebox(title="Erro de Leitura", 
                          message=f"O arquivo de dados está corrompido ou não exite. Será iniciado com uma lista vazia.\n{e}",
                          icon="warning")
            return []
        except Exception as e:
             print(f"Erro ao carregar dados: {e}")
             return []
    else:
        return [] # Inicia vazio se o arquivo não existir
    
def carregar_dados_arquivo(nome_arquivo, label):
    if os.path.exists(nome_arquivo):
        try:
            with open(nome_arquivo, 'r', encoding='utf-8') as f:
                dados_arquivo = json.load(f)
                label.configure(text=dados_arquivo)
        except Exception as e:
            CTkMessagebox(title="Erro de Leitura", 
                          message=f"O arquivo de dados está corrompido ou não exite. Será iniciado com uma lista vazia.\n{e}",
                          icon="warning")
            return []
        except Exception as e:
             print(f"Erro ao carregar dados: {e}")
             return []
    else:
        return [] # Inicia vazio se o arquivo não existir
    
def processar_excel_escala(caminho_escala):
    try:
        escala_df = pd.read_excel(caminho_escala, header=(0))
        colunas_data_objetos = escala_df.columns[3:]
        coluna_turnos = escala_df.columns[1]
        lista_turnos = escala_df[coluna_turnos].dropna().unique().tolist()
        dados_escala = {}

        for coluna_original in colunas_data_objetos:
            if isinstance(coluna_original, pd.Timestamp):
                chave_json = coluna_original.strftime("%Y-%m-%d")
            else:
                chave_json = str(coluna_original)
            nome_por_turno = {}
            
            for turno in lista_turnos:
                condicao_turno = escala_df[coluna_turnos] == turno
                condicao_data = escala_df[coluna_original] == 'OK'
                nomes_encontrados = escala_df.loc[condicao_turno & condicao_data, 'NOME']
                nome_por_turno[turno] = nomes_encontrados.tolist()
            
            dados_escala[chave_json] = nome_por_turno
        
        return dados_escala
    except Exception as e:
        print(f"Erro ao processar Excel: {e}")
        return None
    
